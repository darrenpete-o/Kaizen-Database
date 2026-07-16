import pyodbc
import pandas as pd
from datetime import datetime
import hashlib
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import warnings
warnings.filterwarnings('ignore')

class DatabaseMonitor:
    def __init__(self, baseline_file='database_baseline.json', 
                 excel_file='database_changes.xlsx'):
        """
        Initialize the Database Monitor
        
        Args:
            baseline_file: JSON file to store baseline snapshot
            excel_file: Excel file to document changes
        """
        # Get connection string from environment variable
        self.connection_string = os.environ.get('DB_CONNECTION_STRING')
        if not self.connection_string:
            raise ValueError("DB_CONNECTION_STRING environment variable not set")
        
        self.baseline_file = baseline_file
        self.excel_file = excel_file
        self.conn = None
        self.cursor = None
        
        # Required audit columns
        self.required_columns = ['dtinsert', 'dtedit']
        
    def connect(self):
        """Establish connection to MS SQL Server"""
        try:
            self.conn = pyodbc.connect(self.connection_string)
            self.cursor = self.conn.cursor()
            print("[OK] Connected to database successfully")
            return True
        except Exception as e:
            print(f"[ERROR] Failed to connect to database: {e}")
            return False
    
    def disconnect(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()
            print("[OK] Database connection closed")
    
    def get_database_schema(self):
        """Get current database schema including tables and columns"""
        schema = {}
        
        # First, get all tables with their metadata
        tables_query = """
        SELECT 
            t.TABLE_SCHEMA,
            t.TABLE_NAME,
            -- Get BusinessFriendlyName extended property
            (SELECT TOP 1 CAST(value AS NVARCHAR(MAX)) 
             FROM sys.extended_properties ep
             WHERE ep.major_id = OBJECT_ID(QUOTENAME(t.TABLE_SCHEMA) + '.' + QUOTENAME(t.TABLE_NAME))
               AND ep.minor_id = 0
               AND ep.name = 'BusinessFriendlyName'
               AND ep.class = 1) AS BusinessFriendlyName,
            -- Get FriendlyName extended property
            (SELECT TOP 1 CAST(value AS NVARCHAR(MAX)) 
             FROM sys.extended_properties ep
             WHERE ep.major_id = OBJECT_ID(QUOTENAME(t.TABLE_SCHEMA) + '.' + QUOTENAME(t.TABLE_NAME))
               AND ep.minor_id = 0
               AND ep.name = 'FriendlyName'
               AND ep.class = 1) AS FriendlyName,
            -- Get MS_Description extended property
            (SELECT TOP 1 CAST(value AS NVARCHAR(MAX)) 
             FROM sys.extended_properties ep
             WHERE ep.major_id = OBJECT_ID(QUOTENAME(t.TABLE_SCHEMA) + '.' + QUOTENAME(t.TABLE_NAME))
               AND ep.minor_id = 0
               AND ep.name = 'MS_Description'
               AND ep.class = 1) AS MS_Description,
            -- Get synonym if exists
            (SELECT TOP 1 s.name
             FROM sys.synonyms s
             WHERE OBJECT_ID(s.base_object_name) = OBJECT_ID(QUOTENAME(t.TABLE_SCHEMA) + '.' + QUOTENAME(t.TABLE_NAME))
               AND s.base_object_name IS NOT NULL) AS SynonymName
        FROM 
            INFORMATION_SCHEMA.TABLES t
        WHERE 
            t.TABLE_TYPE = 'BASE TABLE'
            AND t.TABLE_SCHEMA NOT IN ('sys', 'INFORMATION_SCHEMA')
        ORDER BY 
            t.TABLE_SCHEMA, t.TABLE_NAME
        """
        
        self.cursor.execute(tables_query)
        tables = self.cursor.fetchall()
        
        for table in tables:
            schema_name = table[0]
            table_name = table[1]
            business_friendly_name = table[2] if len(table) > 2 else None
            friendly_name = table[3] if len(table) > 3 else None
            ms_description = table[4] if len(table) > 4 else None
            synonym_name = table[5] if len(table) > 5 else None
            
            full_name = f"{schema_name}.{table_name}"
            
            # Determine the best business name and its source
            business_name = None
            name_source = None
            
            if business_friendly_name and str(business_friendly_name).strip():
                business_name = business_friendly_name
                name_source = 'ExtendedProperty_BusinessFriendlyName'
            elif friendly_name and str(friendly_name).strip():
                business_name = friendly_name
                name_source = 'ExtendedProperty_FriendlyName'
            elif ms_description and str(ms_description).strip():
                business_name = ms_description
                name_source = 'ExtendedProperty_MS_Description'
            elif synonym_name and str(synonym_name).strip():
                business_name = synonym_name
                name_source = 'Synonym'
            else:
                # Check if table name itself is business-friendly
                # Remove schema prefix for checking
                clean_name = table_name
                
                # Check for technical prefixes
                technical_prefixes = ['tbl_', 'tab_', 'v_', 'vw_', 'sys_', 'tmp_', 'temp_', 
                                     'lkp_', 'ref_', 'dim_', 'fact_', 'stg_', 'ods_', 'aud_']
                
                technical_suffixes = ['_tbl', '_tab', '_vw', '_view', '_tmp', '_temp']
                
                name_lower = clean_name.lower()
                is_technical = False
                
                for prefix in technical_prefixes:
                    if name_lower.startswith(prefix):
                        is_technical = True
                        break
                
                if not is_technical:
                    for suffix in technical_suffixes:
                        if name_lower.endswith(suffix):
                            is_technical = True
                            break
                
                # Check if it has underscores and isn't title case
                if '_' in clean_name and not clean_name.replace('_', ' ').istitle():
                    is_technical = True
                
                # If it's all lowercase, it's technical
                if name_lower == clean_name:
                    is_technical = True
                
                # If not technical, it's self-descriptive
                if not is_technical:
                    business_name = clean_name.replace('_', ' ')
                    name_source = 'Self-Descriptive'
            
            # Get columns for each table
            columns_query = """
            SELECT 
                COLUMN_NAME,
                DATA_TYPE,
                IS_NULLABLE,
                COLUMN_DEFAULT
            FROM 
                INFORMATION_SCHEMA.COLUMNS
            WHERE 
                TABLE_SCHEMA = ?
                AND TABLE_NAME = ?
            ORDER BY 
                ORDINAL_POSITION
            """
            
            self.cursor.execute(columns_query, (schema_name, table_name))
            columns = self.cursor.fetchall()
            
            schema[full_name] = {
                'business_name': business_name,
                'name_source': name_source,
                'columns': [
                    {
                        'name': col[0],
                        'data_type': col[1],
                        'is_nullable': col[2],
                        'default': col[3]
                    }
                    for col in columns
                ],
                'column_names': [col[0] for col in columns]
            }
        
        return schema
    
    def generate_schema_hash(self, schema):
        """Generate a hash of the schema for change detection"""
        schema_str = json.dumps(schema, sort_keys=True)
        return hashlib.md5(schema_str.encode()).hexdigest()
    
    def load_baseline(self):
        """Load baseline from file if it exists"""
        if os.path.exists(self.baseline_file):
            with open(self.baseline_file, 'r') as f:
                return json.load(f)
        return None
    
    def save_baseline(self, schema, schema_hash):
        """Save current schema as baseline"""
        baseline = {
            'timestamp': datetime.now().isoformat(),
            'schema': schema,
            'hash': schema_hash
        }
        with open(self.baseline_file, 'w') as f:
            json.dump(baseline, f, indent=2)
        print(f"[OK] Baseline saved to {self.baseline_file}")
        return baseline
    
    def detect_changes(self, current_schema, baseline_schema):
        """Detect changes between current and baseline schema"""
        changes = {
            'added_tables': [],
            'removed_tables': [],
            'modified_tables': []
        }
        
        current_tables = set(current_schema.keys())
        baseline_tables = set(baseline_schema.keys())
        
        # Find added tables
        changes['added_tables'] = list(current_tables - baseline_tables)
        
        # Find removed tables
        changes['removed_tables'] = list(baseline_tables - current_tables)
        
        # Check for modifications in existing tables
        common_tables = current_tables & baseline_tables
        for table in common_tables:
            current_cols = set(current_schema[table]['column_names'])
            baseline_cols = set(baseline_schema[table]['column_names'])
            
            if current_cols != baseline_cols:
                changes['modified_tables'].append({
                    'table': table,
                    'added_columns': list(current_cols - baseline_cols),
                    'removed_columns': list(baseline_cols - current_cols)
                })
        
        return changes
    
    def has_business_friendly_name(self, business_name):
        """
        Check if a table has a business friendly name.
        Returns True if ANY business name metadata exists.
        """
        if business_name is None or str(business_name).strip() == '':
            return False
        return True
    
    def create_excel_report(self, changes, current_schema):
        """Create Excel report with changes and warnings"""
        report_data = []
        
        # Process added tables
        for table in changes.get('added_tables', []):
            business_name = current_schema[table].get('business_name')
            name_source = current_schema[table].get('name_source')
            has_business_name = self.has_business_friendly_name(business_name)
            
            # Check for required columns
            has_dtinsert = 'dtinsert' in current_schema[table]['column_names']
            has_dtedit = 'dtedit' in current_schema[table]['column_names']
            
            warning = []
            if not has_business_name:
                warning.append('Missing Business Friendly Name')
            if not has_dtinsert:
                warning.append('Missing dtinsert column')
            if not has_dtedit:
                warning.append('Missing dtedit column')
            
            # Display business name value (or None if missing)
            display_business_name = business_name if business_name else 'None'
            display_source = name_source if name_source else 'None'
            
            # Add table row
            report_data.append({
                'Table Name': table,
                'Business Name': display_business_name,
                'Name Source': display_source,
                'Change Type': 'TABLE ADDED',
                'Column Name': 'TABLE',
                'Data Type': '',
                'Is Nullable': '',
                'Default Value': '',
                'Warnings': '; '.join(warning) if warning else 'None'
            })
            
            # Add column details for new tables
            for col in current_schema[table]['columns']:
                report_data.append({
                    'Table Name': table,
                    'Business Name': display_business_name,
                    'Name Source': display_source,
                    'Change Type': 'COLUMN ADDED',
                    'Column Name': col['name'],
                    'Data Type': col['data_type'],
                    'Is Nullable': col['is_nullable'],
                    'Default Value': str(col['default']) if col['default'] else '',
                    'Warnings': ''
                })
        
        # Process modified tables
        for modification in changes.get('modified_tables', []):
            table = modification['table']
            business_name = current_schema[table].get('business_name')
            name_source = current_schema[table].get('name_source')
            has_business_name = self.has_business_friendly_name(business_name)
            
            display_business_name = business_name if business_name else 'None'
            display_source = name_source if name_source else 'None'
            
            # Added columns
            for col in modification.get('added_columns', []):
                # Find column details from current schema
                col_details = next((c for c in current_schema[table]['columns'] if c['name'] == col), None)
                warning = 'Missing Business Friendly Name' if not has_business_name else ''
                
                report_data.append({
                    'Table Name': table,
                    'Business Name': display_business_name,
                    'Name Source': display_source,
                    'Change Type': 'COLUMN ADDED',
                    'Column Name': col,
                    'Data Type': col_details['data_type'] if col_details else '',
                    'Is Nullable': col_details['is_nullable'] if col_details else '',
                    'Default Value': str(col_details['default']) if col_details and col_details['default'] else '',
                    'Warnings': warning
                })
            
            # Removed columns
            for col in modification.get('removed_columns', []):
                report_data.append({
                    'Table Name': table,
                    'Business Name': display_business_name,
                    'Name Source': display_source,
                    'Change Type': 'COLUMN REMOVED',
                    'Column Name': col,
                    'Data Type': '',
                    'Is Nullable': '',
                    'Default Value': '',
                    'Warnings': 'Column removed from table'
                })
        
        # Process removed tables
        for table in changes.get('removed_tables', []):
            report_data.append({
                'Table Name': table,
                'Business Name': 'Unknown (Table Removed)',
                'Name Source': 'Unknown',
                'Change Type': 'TABLE REMOVED',
                'Column Name': 'TABLE',
                'Data Type': '',
                'Is Nullable': '',
                'Default Value': '',
                'Warnings': 'Table no longer exists in database'
            })
        
        # If no changes detected, return None
        if not report_data:
            print("[OK] No changes detected")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(report_data)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{self.excel_file.replace('.xlsx', '')}_{timestamp}.xlsx"
        
        # Write to Excel with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Changes', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Changes']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply formatting for warnings
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            # Highlight warning cells (column 8 in the Excel file)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=8, max_col=8):
                for cell in row:
                    if cell.value and cell.value != 'None':
                        if 'Missing Business Friendly Name' in str(cell.value):
                            cell.fill = yellow_fill
                            cell.font = Font(bold=True)
                        if 'Missing dtinsert' in str(cell.value) or 'Missing dtedit' in str(cell.value):
                            cell.fill = red_fill
                            cell.font = Font(bold=True, color='FFFFFF')
        
        print(f"[OK] Report saved to {filename}")
        return filename
    
    def run_monitoring(self):
        """Main monitoring function"""
        print("[INFO] Starting database monitoring...")
        
        if not self.connect():
            return False
        
        try:
            # Get current schema
            current_schema = self.get_database_schema()
            current_hash = self.generate_schema_hash(current_schema)
            
            # Load baseline
            baseline = self.load_baseline()
            
            if baseline is None:
                print("[INFO] No baseline found. Creating initial baseline...")
                self.save_baseline(current_schema, current_hash)
                print("[OK] Baseline created successfully.")
                return True
            
            # Check if changes occurred
            if baseline['hash'] == current_hash:
                print("[OK] No changes detected in the database.")
                return True
            
            # Detect and document changes
            print("[INFO] Changes detected! Generating report...")
            changes = self.detect_changes(current_schema, baseline['schema'])
            
            # Create Excel report
            report_file = self.create_excel_report(changes, current_schema)
            
            # Update baseline
            self.save_baseline(current_schema, current_hash)
            
            # Print summary
            print("\n[SUMMARY] Change Summary:")
            print(f"  - Added tables: {len(changes['added_tables'])}")
            print(f"  - Removed tables: {len(changes['removed_tables'])}")
            print(f"  - Modified tables: {len(changes['modified_tables'])}")
            if report_file:
                print(f"  - Report saved to: {report_file}")
            
            return True
            
        except Exception as e:
            print(f"[ERROR] Error during monitoring: {e}")
            return False
        finally:
            self.disconnect()

# Example usage
if __name__ == "__main__":
    # Initialize and run monitor
    monitor = DatabaseMonitor(
        baseline_file='database_baseline.json',
        excel_file='database_changes.xlsx'
    )
    
    success = monitor.run_monitoring()
    
    # Exit with appropriate code for CI/CD
    exit(0 if success else 1)
